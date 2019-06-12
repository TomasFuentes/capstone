from openpyxl import load_workbook, utils
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import json

def cargar_base(archivo, termino, nombre_pestana):
    wb = load_workbook(archivo)
    wsdatos = wb[nombre_pestana]
    calles = {}
    for x in range(4,termino+1):
        nombrecalle= wsdatos.cell(row=x, column=1).value
        dicc_aux = {}
        comienza = (wsdatos.cell(row=x, column=2).value, wsdatos.cell(row=x, column=3).value)
        termina = (wsdatos.cell(row=x, column=4).value, wsdatos.cell(row=x, column=5).value)
        dicc_aux["basura_max"]= wsdatos.cell(row=x, column=6).value
        dicc_aux["comienza"] = comienza
        dicc_aux["termina"] = termina
        calles[nombrecalle] = dicc_aux
    return calles


def datos_historicos(archivo, termino, nombre_pestana,diccionario_calles):
    wb = load_workbook(archivo)
    wsdatos = wb[nombre_pestana]
    lista_datos = []
    dic_final = diccionario_calles
    for x in range(2,termino+1):
        nombrecalle = wsdatos.cell(row=x, column=2).value
        lista_aux = (wsdatos.cell(row=x, column=1).value, nombrecalle, wsdatos.cell(row=x, column=3).value)
        lista_datos.append(lista_aux)
    calles = []
    for x in range(1,6111):
        calles.append("Calle " + str(x))
    for calle in calles:
        datos_calle =[]
        for tupla in lista_datos:
            if tupla[1] == calle:
                datos_calle.append(tupla[2])
        dic_final[calle]["datos_historicos"] = datos_calle
        dic_final[calle]["esperanza"] = np.mean(datos_calle)
    return dic_final

datos = cargar_base("Datos.xlsx",6113,"Ciudad")
datos2 = datos_historicos("Datos.xlsx",171081,"Histórico de basura",datos)

with open('data.txt', 'w') as outfile:
    json.dump(datos2, outfile)